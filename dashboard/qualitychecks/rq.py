import pandas as pd
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def rq(df_adlib):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="output.xlsx"'
    
    df_01 = df_adlib[df_adlib['objectnaam'].isna()] 
    df_02 = df_adlib[df_adlib['titel'].isna()] 
    df_03 = df_adlib[df_adlib['reproductie.referentie'].isna()] 
    df_04 = df_adlib[df_adlib['associatie.onderwerp'].isna()]
    df_05 = df_adlib[df_adlib["instelling.naam"] != 'Het Huis van Alijn (Gent)']

    df_06 = df_adlib[df_adlib['vervaardiging.datum.begin'].notna()]
    df_06 = df_06[df_06['associatie.periode'].isna()]

    df_07 = df_adlib[df_adlib["titel"].str.contains("Gent", na=False)]
    df_07 = df_07[~df_07["associatie.onderwerp"].str.contains("Gent", na=False)]

    df_08 = df_adlib[df_adlib['onderscheidende_kenmerken']!='DIGITALE COLLECTIE']
    df_08 = df_08[df_08['onderscheidende_kenmerken']!='OBJECT']
    df_08 = df_08[df_08['onderscheidende_kenmerken']!='BEELD']
    df_08 = df_08[df_08['onderscheidende_kenmerken']!='DOCUMENTAIRE COLLECTIE']
    df_08 = df_08[df_08['onderscheidende_kenmerken']!='TEXTIEL']
    df_08 = df_08[df_08['onderscheidende_kenmerken']!='AUDIOVISUELE COLLECTIE']

    df_09 = df_adlib[df_adlib["vervaardiging.plaats"].str.contains("Gent", na=False)]
    df_09 = df_09[~df_09["associatie.onderwerp"].str.contains("Gent", na=False)]

    wb = Workbook()
    ws = wb.active
    ws.title = 'Info'
    ws['A1'] = "list of sheet tab codes"
    ws.append(['sheet number', 'quality check'])
    ws.append(['#01', 'missing object_name'])
    ws.append(['#02', 'missing title'])
    ws.append(['#03', 'missing images'])
    ws.append(['#04', 'missing associated_subject'])
    ws.append(['#05', 'wrong institution name'])
    ws.append(['#06', 'missing associated_period'])
    ws.append(['#07', 'title = Gent AND NOT associated_subject = Gent'])
    ws.append(['#08', 'missing or wrong onderscheidende_kenmerken'])
    ws.append(['#09', 'place creation = Gent AND NOT associated_subject = Gent'])
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 60

    if df_01.empty == True:
        print('empty dataframe')
    else:
        ws = wb.create_sheet("#01")
        rows = dataframe_to_rows(df_01, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    if df_02.empty == True:
        print('empty dataframe')
    else:
        ws = wb.create_sheet("#02")
        rows = dataframe_to_rows(df_02, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    if df_03.empty == True:
        print('empty dataframe')
    else:
        ws = wb.create_sheet("#03")
        rows = dataframe_to_rows(df_03, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    if df_04.empty == True:
        print('empty dataframe')
    else:
        ws = wb.create_sheet("#04")
        rows = dataframe_to_rows(df_04, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    if df_05.empty == True:
        print('empty dataframe')
    else:
        ws = wb.create_sheet("#05")
        rows = dataframe_to_rows(df_05, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    if df_06.empty == True:
        print('empty dataframe')
    else:
        ws = wb.create_sheet("#06")
        rows = dataframe_to_rows(df_06, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    if df_07.empty == True:
        print('empty dataframe')
    else:
        ws = wb.create_sheet("#07")
        rows = dataframe_to_rows(df_07, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    if df_08.empty == True:
        print('empty dataframe')
    else:
        ws = wb.create_sheet("#08")
        rows = dataframe_to_rows(df_08, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    if df_09.empty == True:
        print('empty dataframe')
    else:
        ws = wb.create_sheet("#09")
        rows = dataframe_to_rows(df_09, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response