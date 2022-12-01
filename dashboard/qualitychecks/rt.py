import pandas as pd
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def rt(df_thesaurus):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="output.xlsx"'
    
    df_01 = df_thesaurus[df_thesaurus['bron']!='http://vocab.getty.edu/aat/']
    df_01 = df_01[df_01['bron']!='https://id.erfgoed.net/themas/']
    df_01 = df_01[df_01['bron']!='http://vocab.getty.edu/tgn/']
    df_01 = df_01[df_01['bron']!='https://www.wikidata.org/entity/']
    df_01 = df_01[df_01['bron']!='https://id.erfgoed.net/erfgoedobjecten/']

    df_02 = df_thesaurus['term'].value_counts()
    df_02 = df_02.loc[lambda x : x > 1]
    df_02 = pd.DataFrame({'term':df_02.index, 'number of occurences':df_02.values})

    df_03 = df_thesaurus['term.nummer'].value_counts()
    df_03 = df_03.loc[lambda x : x > 1]
    df_03 = pd.DataFrame({'term':df_03.index, 'number of occurences':df_03.values})

    df_04 = df_thesaurus[df_thesaurus['bron']=='https://www.wikidata.org/entity/']
    df_04 = df_04[~df_04["term.nummer"].str.contains("Q", na=False)]

    df_05 = df_thesaurus[df_thesaurus['bron'] == 'http://vocab.getty.edu/aat/']
    df_05 = df_05[~df_05['term.nummer'].isna()]
    df_05 = df_05['term.nummer'].astype(int)
    df_05 = pd.DataFrame({'term':df_05.values})
    df_05 = df_05[(df_05['term']>999999999) | (df_05['term']<100000000)]

    df_06 = df_thesaurus[df_thesaurus['bron'] == 'http://vocab.getty.edu/tgn/']
    df_06 = df_06[~df_06['term.nummer'].isna()]
    df_06 = df_06['term.nummer'].astype(int)
    df_06 = pd.DataFrame({'term':df_06.values})
    df_06 = df_06[(df_06['term']>9999999) | (df_06['term']<1000000)]

    wb = Workbook()
    ws = wb.active
    ws.title = 'Info'
    ws['A1'] = "list of sheet tab codes"
    ws.append(['sheet number', 'quality check'])
    ws.append(['#01', 'missing or wrong external authority'])
    ws.append(['#02', 'double terms'])
    ws.append(['#03', 'double external authorities'])
    ws.append(['#04', 'wrong Wikidata-number'])
    ws.append(['#05', 'wrong AAT-number'])
    ws.append(['#06', 'wrong TGN-number'])
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 60

    if df_01.empty == True:
        print('empty dataframe')
    else:
        ws = wb.create_sheet("#01")
        rows = dataframe_to_rows(df_01, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    if df_02.empty == True:
        print('empty dataframe')
    else:
        ws = wb.create_sheet("#02")
        rows = dataframe_to_rows(df_02, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    if df_03.empty == True:
        print('empty dataframe')
    else:
        ws = wb.create_sheet("#03")
        rows = dataframe_to_rows(df_03, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    if df_04.empty == True:
        print('empty dataframe')
    else:
        ws = wb.create_sheet("#04")
        rows = dataframe_to_rows(df_04, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    if df_05.empty == True:
        print('empty dataframe')
    else:
        ws = wb.create_sheet("#05")
        rows = dataframe_to_rows(df_05, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    if df_06.empty == True:
        print('empty dataframe')
    else:
        ws = wb.create_sheet("#06")
        rows = dataframe_to_rows(df_06, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response